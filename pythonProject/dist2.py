import pandas as pd
from openpyxl.styles import PatternFill

def load_data(file_path):
    return pd.read_excel(file_path)

def check_duplicate_neptun_codes(data):
    duplicated = data[data.duplicated(subset=['Neptun kód'], keep=False)]
    if not duplicated.empty:
        print("Warning: There are students with duplicate Neptun kód:")
        for neptun_code in duplicated['Neptun kód'].unique():
            print(f"Neptun kód: {neptun_code}")
    else:
        print("No duplicate Neptun kód found.")

def group_students(data):
    bins = [0, 2, 4, 6, 8, 10, 12]
    labels = ['1. éves', '2. éves', '3. éves', '4. éves', '5. éves', '6. éves']
    data['Évfolyam'] = pd.cut(data['Aktív félévek'], bins=bins, labels=labels, right=True)

    grouping_columns = ['KépzésNév', 'Képzési szint', 'Nyelv ID', 'Évfolyam']
    grouped = data.groupby(grouping_columns, observed=True).size().reset_index(name='Létszám')
    return grouped, data

def find_nearest_group(group, i):
    prev_idx, next_idx = i - 1, i + 1
    closest_idx = None
    while prev_idx >= 0 or next_idx < len(group):
        if prev_idx >= 0 and group.iloc[prev_idx]['Létszám'] >= 10:
            closest_idx = prev_idx if closest_idx is None else (
                prev_idx if group.iloc[prev_idx]['Létszám'] < group.iloc[closest_idx]['Létszám'] else closest_idx)
        if next_idx < len(group) and group.iloc[next_idx]['Létszám'] >= 10:
            closest_idx = next_idx if closest_idx is None else (
                next_idx if group.iloc[next_idx]['Létszám'] < group.iloc[closest_idx]['Létszám'] else closest_idx)
        if closest_idx is not None:
            return closest_idx
        prev_idx -= 1
        next_idx += 1
    return None

def redistribute_students(grouped, original_data):
    modified_data = original_data.copy()

    grouping_columns = ['KépzésNév', 'Képzési szint', 'Nyelv ID']

    for keys in grouped[grouping_columns].drop_duplicates().values:
        key_dict = dict(zip(grouping_columns, keys))
        group_mask = (grouped['KépzésNév'] == keys[0]) & \
                     (grouped['Képzési szint'] == keys[1]) & \
                     (grouped['Nyelv ID'] == keys[2])
        group = grouped[group_mask].copy().reset_index(drop=True)

        total_students = group['Létszám'].sum()
        if total_students < 10:
            continue

        for i in range(len(group)):
            if group.iloc[i]['Létszám'] < 10 and group.iloc[i]['Létszám'] > 0:
                current_year = group.iloc[i]['Évfolyam']
                original_index = modified_data[
                    (modified_data['KépzésNév'] == keys[0]) &
                    (modified_data['Képzési szint'] == keys[1]) &
                    (modified_data['Nyelv ID'] == keys[2]) &
                    (modified_data['Évfolyam'] == current_year)].index

                closest_year_index = find_nearest_group(group, i)

                if closest_year_index is not None:
                    new_year = group.iloc[closest_year_index]['Évfolyam']
                    modified_data.loc[original_index, 'Évfolyam'] = new_year
                    group.at[closest_year_index, 'Létszám'] += group.at[i, 'Létszám']
                    group.at[i, 'Létszám'] = 0

    return modified_data

def calculate_scholarship_index(data):
    data['Kredit szám'] = data['ElőzőFélévTeljesítettKredit'].apply(lambda x: min(x, 42))
    data['Ösztöndíjindex'] = data['Ösztöndíj átlag előző félév'] + ((data['Kredit szám'] / 27) - 1) / 2
    return data

def recalculate_year_for_small_groups(data):
    bins = [0, 2, 4, 6, 8, 10, 12]
    labels = ['1. éves', '2. éves', '3. éves', '4. éves', '5. éves', '6. éves']
    data['Évfolyam'] = pd.cut(data['Aktív félévek'], bins=bins, labels=labels, right=True)
    return data

def save_to_excel(main_data, separate_data, output_file, separate_file):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        main_data.to_excel(writer, index=False, sheet_name='MainData')
        apply_alternate_row_coloring(writer, main_data, 'MainData')

    with pd.ExcelWriter(separate_file, engine='openpyxl') as writer:
        separate_data.to_excel(writer, index=False, sheet_name='SeparateData')
        apply_alternate_row_coloring(writer, separate_data, 'SeparateData')

def apply_alternate_row_coloring(writer, df, sheet_name):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    grouping_columns = ['KépzésNév', 'Képzési szint', 'Nyelv ID', 'Évfolyam']
    df['GroupID'] = df[grouping_columns].apply(lambda x: ' | '.join(x.astype(str)), axis=1)
    last_row = df.shape[0] + 1
    last_col = df.shape[1]

    fill_colors = ['FFFFFF', 'D3D3D3']
    current_fill = 0
    previous_group = None

    for row in range(2, last_row + 1):
        group_id = df.iloc[row - 2]['GroupID']
        if group_id != previous_group:
            current_fill = (current_fill + 1) % len(fill_colors)
            previous_group = group_id

        fill = PatternFill(start_color=fill_colors[current_fill], end_color=fill_colors[current_fill], fill_type='solid')
        for col in range(1, last_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill

    df.drop(columns=['GroupID'], inplace=True)

def filter_small_groups(data):
    grouping_columns = ['KépzésNév', 'Képzési szint', 'Nyelv ID']
    course_counts = data.groupby(grouping_columns).size().reset_index(name='Total_in_Course')
    data = data.merge(course_counts, on=grouping_columns)
    small_groups = data[data['Total_in_Course'] < 10]
    remaining_data = data[data['Total_in_Course'] >= 10]
    small_groups = small_groups.drop(columns=['Total_in_Course'])
    remaining_data = remaining_data.drop(columns=['Total_in_Course'])
    return remaining_data, small_groups

def sort_data(data):
    data = data.sort_values(by=['KépzésNév', 'Képzési szint', 'Nyelv ID', 'Évfolyam', 'Aktív félévek'], ascending=True)
    return data

def calculate_kodi(data):
    grouping_columns = ['KépzésNév', 'Képzési szint', 'Nyelv ID', 'Évfolyam']
    grouped = data.groupby(grouping_columns, observed=True)

    data['MinÖDI'] = grouped['Ösztöndíjindex'].transform('min')
    data['MaxÖDI'] = grouped['Ösztöndíjindex'].transform('max')

    def calculate_group_kodi(row):
        if row['MaxÖDI'] == row['MinÖDI']:
            return 100.0
        else:
            kodi = ((row['Ösztöndíjindex'] - row['MinÖDI']) / (row['MaxÖDI'] - row['MinÖDI'])) * 100
            return round(kodi, 6)

    data['KÖDI'] = data.apply(calculate_group_kodi, axis=1)

    max_odi_mask = data['Ösztöndíjindex'] == data['MaxÖDI']
    data.loc[max_odi_mask, 'KÖDI'] = 100.0

    min_odi_mask = data['Ösztöndíjindex'] == data['MinÖDI']
    data.loc[min_odi_mask, 'KÖDI'] = 0.0

    data.drop(columns=['MinÖDI', 'MaxÖDI'], inplace=True)

    return data

def remove_lower_kodi_duplicates(data):
    duplicated = data[data.duplicated(subset=['Neptun kód'], keep=False)]
    if not duplicated.empty:
        print("Removing lower KÖDI entries for duplicate Neptun kód:")
        for neptun_code in duplicated['Neptun kód'].unique():
            student_rows = data[data['Neptun kód'] == neptun_code]
            max_kodi_index = student_rows['KÖDI'].idxmax()
            indices_to_drop = student_rows.index.difference([max_kodi_index])
            data = data.drop(indices_to_drop)
            print(f"Neptun kód {neptun_code}: Kept index {max_kodi_index}, dropped indices {list(indices_to_drop)}")
    else:
        print("No duplicate Neptun kód found after KÖDI calculation.")
    return data.reset_index(drop=True)

# Input and Output file paths
input_file = '/Users/bercelkovalik/Documents./InputOutput/Adatok.xlsx'
output_file = '/Users/bercelkovalik/Documents./InputOutput/output_data.xlsx'
separate_file = '/Users/bercelkovalik/Documents./InputOutput/small_groups_output.xlsx'

# Load data
data = load_data(input_file)

# Check for duplicate Neptun kód entries before processing
check_duplicate_neptun_codes(data)

# Initial Function Calls
remaining_data, small_groups_data = filter_small_groups(data)
grouped_data, original_data = group_students(remaining_data)
updated_data = redistribute_students(grouped_data, original_data)
updated_data = calculate_scholarship_index(updated_data)
updated_data = calculate_kodi(updated_data)

updated_data = remove_lower_kodi_duplicates(updated_data)

# Re-run the grouping and redistribution after removing duplicates
# **This step addresses the group size changes due to eliminations**
remaining_data, small_groups_data = filter_small_groups(updated_data)
grouped_data, original_data = group_students(remaining_data)
updated_data = redistribute_students(grouped_data, original_data)
updated_data = calculate_scholarship_index(updated_data)
updated_data = calculate_kodi(updated_data)

# Sorting
updated_data = sort_data(updated_data)
small_groups_data = sort_data(small_groups_data)

# Save to Excel
save_to_excel(updated_data, small_groups_data, output_file, separate_file)

print("Process completed.")
