import pandas as pd
from openpyxl.styles import PatternFill
import streamlit as st
import io
"""
STEP 1
Legtöbb függvényhez írtam help tag description-t hogy könnyebb legyen értelmezni
"""
###Copyright 2024, Bercel Kovalik-Deák, All rights reserved
@st.cache_data

### Excel betöltés
def load_data(file_path):
    """
    Betölti az adatokat egy Excel fájlból.

        Args:
            file_path (str): Az Excel fájl elérési útja.

        Returns:
            pd.DataFrame: A betöltött adatok DataFrame-ként.
            Raises:
                FileNotFoundError: Ha a fájl nem található.
                pd.errors.ParserError: Ha az Excel fájl nem megfelelően formázott.
    """
    return pd.read_excel(file_path)

### Megnézi van két vagy több képzést végző hallgató, ha igen akkor kiírja a Neptun kódját a hallgatónak
def check_duplicate_neptun_codes(data):
    """Ellenőrzi, hogy vannak-e duplikált Neptun kódok az adatokban.

        Args:
            data (pd.DataFrame): A hallgatók adatait tartalmazó DataFrame.
        """
    duplicated = data[data.duplicated(subset=['Neptun kód'], keep=False)]
    if not duplicated.empty:
        print("Warning: There are students with duplicate Neptun kód:")
        for neptun_code in duplicated['Neptun kód'].unique():
            print(f"Neptun kód: {neptun_code}")
    else:
        print("No duplicate Neptun kód found.")

### Folytatása az előző functionnek, ha talált többször szereplő neptun kódot, akkor kitörli azt a sort amiben a
### hallgatónak alacsonyabb KÖDI értéke van.
def remove_lower_kodi_duplicates(data):
    """Eltávolítja az alacsonyabb KÖDI értékkel rendelkező duplikált Neptun kódokat.

        Args:
            data (pd.DataFrame): A hallgatók adatait tartalmazó DataFrame.

        Returns:
            pd.DataFrame: A DataFrame a duplikált bejegyzések eltávolítása után.
        """
    duplicated = data[data.duplicated(subset=['Neptun kód'], keep=False)]
    if not duplicated.empty:
        print("Removing lower KÖDI entries for duplicate Neptun kód:")
        for neptun_code in duplicated['Neptun kód'].unique():
            student_rows = data[data['Neptun kód'] == neptun_code]
            max_kodi_index = student_rows['KÖDI'].idxmax()
            indices_to_drop = student_rows.index.difference([max_kodi_index])
            data = data.drop(indices_to_drop)
            print(f"Neptun kód {neptun_code}: Kept index {max_kodi_index}, dropped indices {list(indices_to_drop)}")
    else:
        print("No duplicate Neptun kód found after KÖDI calculation.")
    return data.reset_index(drop=True)

### Ellenőrzi hogy van-e olyan hallgató aki túllépte a jogosultsági időszakot a képzés típusa alapján
# (alap, mester, osztatlan) és hozzá ad 1-et az alaphoz és mesterhez, és 2-t az osztatlanhoz.
def highlight_exceeded_semesters(main_data, small_groups_data, semester_limits):
    """Kiemeli azokat a hallgatókat, akik túllépték a jogosultsági időszakot.

        A képzés típusa alapján (alap, mester, osztatlan) hozzáad 1-et az alap- és mesterképzéshez,
        és 2-t az osztatlan képzéshez az aktív félévek számához.

        Args:
            main_data (pd.DataFrame): A fő adatokat tartalmazó DataFrame.
            small_groups_data (pd.DataFrame): A kiscsoportok adatait tartalmazó DataFrame.
            semester_limits (pd.DataFrame): A félévszám limiteket tartalmazó DataFrame.

        Returns:
            tuple: A módosított main_data és small_groups_data DataFrame-ek.
    """
    main_data = pd.merge(main_data, semester_limits, how='left', left_on='KépzésKód', right_on='Képzéskód')
    small_groups_data = pd.merge(small_groups_data, semester_limits, how='left', left_on='KépzésKód', right_on='Képzéskód')

    main_data['Adjusted Félévszám'] = main_data['Félévszám'] + main_data['Képzési szint_x'].map(
        {'alapképzés (BA/BSc/BProf)': 1, 'mesterképzés (MA/MSc)': 1, 'egységes, osztatlan képzés': 2}).fillna(0)
    small_groups_data['Adjusted Félévszám'] = small_groups_data['Félévszám'] + small_groups_data['Képzési szint_x'].map(
        {'alapképzés (BA/BSc/BProf)': 1, 'mesterképzés (MA/MSc)': 1, 'egységes, osztatlan képzés': 2}).fillna(0)

    main_data['Exceed Limit'] = main_data['Aktív félévek'] > main_data['Adjusted Félévszám']
    small_groups_data['Exceed Limit'] = small_groups_data['Aktív félévek'] > small_groups_data['Adjusted Félévszám']

    columns_to_drop = ['Képzéskód', 'Félévszám', 'Képzés neve', 'Képzési szint_y', 'Tagozat_y', 'FIR-be felad', 'Nyelv', 'Szükséges kredit',
                       'Létrehozás ideje', 'Utolsó módosítás ideje', 'Létrehozó', 'Utolsó módosító', 'Archivált',
                       'Szervezeti egység kódja']
    main_data = main_data.drop(columns=[col for col in columns_to_drop if col in main_data.columns])
    small_groups_data = small_groups_data.drop(
        columns=[col for col in columns_to_drop if col in small_groups_data.columns])

    return main_data, small_groups_data

### Itt először csak átírja a félévszámokból szöveges labelre, és rendezi őket 4 tagú kulcs szerint
def group_students(data):
    """Csoportosítja a hallgatókat évfolyam szerint.

    Args:
        data (pd.DataFrame): A hallgatók adatait tartalmazó DataFrame.

    Returns:
        tuple: A csoportosított adatokat és az eredeti adatokat tartalmazó tuple.
    """
    bins = [0, 2, 4, 6, 8, 10, 12]
    labels = ['1. éves', '2. éves', '3. éves', '4. éves', '5. éves', '6. éves']
    data['Évfolyam'] = pd.cut(data['Aktív félévek'], bins=bins, labels=labels, right=True)

    grouping_columns = ['KépzésNév', 'Képzési szint', 'Nyelv ID', 'Évfolyam']
    grouped = data.groupby(grouping_columns, observed=True).size().reset_index(name='Létszám')
    return grouped, data

# Azokon a szakokon ahol nincsen meg összesen a 10 fő, a hallgatókat egy külön excelbe rakja, small_groups.xlsx
def filter_small_groups(data):
    """Kiszűri azokat a hallgatókat, akik kiscsoportba tartoznak (kevesebb mint 10 fő a szakon).

        Args:
            data (pd.DataFrame): A hallgatók adatait tartalmazó DataFrame.

        Returns:
            tuple: A maradék adatokat és a kiscsoportok adatait tartalmazó tuple.
    """
    grouping_columns = ['KépzésNév', 'Képzési szint', 'Nyelv ID']
    course_counts = data.groupby(grouping_columns).size().reset_index(name='Total_in_Course')
    data = data.merge(course_counts, on=grouping_columns)
    small_groups = data[data['Total_in_Course'] < 10]
    remaining_data = data[data['Total_in_Course'] >= 10]
    small_groups = small_groups.drop(columns=['Total_in_Course'])
    remaining_data = remaining_data.drop(columns=['Total_in_Course'])
    return remaining_data, small_groups

### Ez a main logic, jó hosszú és itt történik a legfontosabb dolog, a hallgatók szak szerinti és évfolyam szerinti
# legalább 10 fős csoportokba rendezése. Ha egy évfolyam nincs meg 10 fő, akkor addig iterál jobbra és balra is amíg
# talál mindkét irányba csoportot, ha talált akkor összevonja a kisebb létszámú évfolyammal.
# Így működik, PÉLDA:               1. évfolyam / 2. évfolyam / 3. évfolyam / 4. évfolyam
# Feltételes helyzet, (db hallgató)       12            0           6           51
# Rendezés után:                          18            0           0           51

def group_students_by_year(data):
    """Csoportosítja a hallgatókat évfolyam szerint, biztosítva, hogy minden csoport legalább 10 fős legyen.

        Args:
            data (pd.DataFrame): A hallgatók adatait tartalmazó DataFrame.

        Returns:
            pd.DataFrame: A módosított DataFrame az összevont évfolyamokkal.
    """
    modified_data = data.copy()
    grouping_columns = ['KépzésNév', 'Képzési szint', 'Nyelv ID']

    unique_programs = modified_data[grouping_columns].drop_duplicates()

    for _, program in unique_programs.iterrows():
        program_mask = (
            (modified_data['KépzésNév'] == program['KépzésNév']) &
            (modified_data['Képzési szint'] == program['Képzési szint']) &
            (modified_data['Nyelv ID'] == program['Nyelv ID'])
        )
        program_data = modified_data[program_mask].copy()

        year_counts = program_data['Évfolyam'].value_counts().sort_index()
        years = list(year_counts.index)
        counts = list(year_counts.values)

        year_df = pd.DataFrame({'Évfolyam': years, 'Létszám': counts})
        year_df.sort_values('Évfolyam', inplace=True)
        year_df.reset_index(drop=True, inplace=True)

        year_order = {label: idx for idx, label in enumerate(year_df['Évfolyam'])}

        year_to_group = {year: year for year in year_df['Évfolyam']}

        merged_years = set()

        idx = 0
        while idx < len(year_df):
            current_year = year_df.loc[idx, 'Évfolyam']
            current_count = year_df.loc[idx, 'Létszám']

            if current_count >= 10:
                idx += 1
                continue

            prev_idx = idx - 1 if idx > 0 else None
            next_idx = idx + 1 if idx + 1 < len(year_df) else None

            candidates = []
            if prev_idx is not None:
                prev_year = year_df.loc[prev_idx, 'Évfolyam']
                prev_count = year_df.loc[prev_idx, 'Létszám']
                candidates.append({'idx': prev_idx, 'year': prev_year, 'count': prev_count})
            if next_idx is not None:
                next_year = year_df.loc[next_idx, 'Évfolyam']
                next_count = year_df.loc[next_idx, 'Létszám']
                candidates.append({'idx': next_idx, 'year': next_year, 'count': next_count})

            if not candidates:
                idx += 1
                continue

            min_count = min(c['count'] for c in candidates)
            min_candidates = [c for c in candidates if c['count'] == min_count]

            if len(min_candidates) > 1:
                merge_candidate = max(min_candidates, key=lambda c: year_order[c['year']])
            else:
                merge_candidate = min_candidates[0]

            merge_idx = merge_candidate['idx']
            merge_year = merge_candidate['year']

            year_df.loc[merge_idx, 'Létszám'] += current_count
            year_df.loc[idx, 'Létszám'] = 0

            target_group = year_to_group[merge_year]
            year_to_group[current_year] = target_group

            modified_data.loc[
                program_mask & (modified_data['Évfolyam'] == current_year),
                'Évfolyam'
            ] = target_group

            merged_years.add(current_year)
            year_df = year_df[year_df['Létszám'] > 0].reset_index(drop=True)
            idx = 0

        for merged_year in merged_years:
            target_group = year_to_group[merged_year]
            modified_data.loc[
                program_mask & (modified_data['Évfolyam'] == merged_year),
                'Évfolyam'
            ] = target_group

    return modified_data

# Ösztöndíjindex kiszámolása
def calculate_scholarship_index(data):
    """Kiszámolja az ösztöndíjindexet.

        Args:
            data (pd.DataFrame): A hallgatók adatait tartalmazó DataFrame.

        Returns:
            pd.DataFrame: A DataFrame a kiszámolt ösztöndíjindexszel.
        """
    data['Kredit szám'] = data['ElőzőFélévTeljesítettKredit'].apply(lambda x: min(x, 42))
    data['Ösztöndíjindex'] = data['Ösztöndíj átlag előző félév'] + ((data['Kredit szám'] / 27) - 1) / 2
    return data
# A 10 főnél kisebb szak hallgatóira újra számoljuk az évfolyamot
def recalculate_year_for_small_groups(data):
    """Újraszámolja az évfolyamot a kiscsoportok hallgatói számára.

        Args:
            data (pd.DataFrame): A hallgatók adatait tartalmazó DataFrame.

        Returns:
            pd.DataFrame: A DataFrame az újraszámolt évfolyamokkal.
        """
    bins = [0, 2, 4, 6, 8, 10, 12]
    labels = ['1. éves', '2. éves', '3. éves', '4. éves', '5. éves', '6. éves']
    data['Évfolyam'] = pd.cut(data['Aktív félévek'], bins=bins, labels=labels, right=True)
    return data
# Magától értetődő
def save_to_excel(main_data, separate_data):
    main_buffer = io.BytesIO()
    with pd.ExcelWriter(main_buffer, engine='openpyxl') as writer:
        main_data.to_excel(writer, index=False, sheet_name='MainData')
        apply_alternate_row_coloring(writer, main_data, 'MainData')
    main_buffer.seek(0)


    separate_buffer = io.BytesIO()
    with pd.ExcelWriter(separate_buffer, engine='openpyxl') as writer:
        separate_data.to_excel(writer, index=False, sheet_name='SeparateData')
        apply_alternate_row_coloring(writer, separate_data, 'SeparateData')
    separate_buffer.seek(0)

    return main_buffer, separate_buffer

# Csak hogy könnyebben megkülönböztethetőek legyen a csoportok az output excelben, színezi a sorokat
def apply_alternate_row_coloring(writer, df, sheet_name):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    grouping_columns = ['KépzésNév', 'Képzési szint_x', 'Nyelv ID', 'Évfolyam']
    df['TempGroupID'] = df[grouping_columns].apply(lambda x: ' | '.join(x.astype(str)), axis=1)
    last_row = df.shape[0] + 1
    last_col = df.shape[1]

    fill_colors = ['FFFFFF', 'D3D3D3']
    current_fill = 0
    previous_group = None
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    for row in range(2, last_row + 1):
        group_id = df.iloc[row - 2]['TempGroupID']
        exceed_limit = df.iloc[row - 2]['Exceed Limit'] if 'Exceed Limit' in df.columns else False

        if exceed_limit:
            for col in range(1, last_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = red_fill
        else:
            if group_id != previous_group:
                current_fill = (current_fill + 1) % len(fill_colors)
                previous_group = group_id

            fill = PatternFill(start_color=fill_colors[current_fill], end_color=fill_colors[current_fill],
                               fill_type='solid')
            for col in range(1, last_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = fill

    df.drop(columns=['TempGroupID'], inplace=True)

# Beszámozza a csoportokat
def add_group_index(data):
    """Hozzáad egy csoportindexet az adatokhoz.

        Args:
            data (pd.DataFrame): A hallgatók adatait tartalmazó DataFrame.

        Returns:
            pd.DataFrame: A DataFrame a hozzáadott csoportindexszel.
        """
    grouping_columns = ['KépzésNév', 'Képzési szint', 'Nyelv ID', 'Évfolyam']
    data['GroupID'] = data[grouping_columns].apply(lambda x: ' | '.join(x.astype(str)), axis=1)
    data['GroupIndex'] = (data['GroupID'] != data['GroupID'].shift()).cumsum()
    data.drop(columns=['GroupID'], inplace=True)
    cols = data.columns.tolist()
    cols.insert(0, cols.pop(cols.index('GroupIndex')))
    data = data[cols]
    return data

# Rendezés
def sort_data(data):
    data = data.sort_values(by=['KépzésNév', 'Képzési szint', 'Nyelv ID', 'Évfolyam', 'Aktív félévek'], ascending=True)
    return data
# KÖDI számolás, minden csoportban a max Ösztöndíjindex kap egy 100-as értéket, a legalacsonyabb pedig 0-t.
# És a csoport többi tagja hozzájuk aránylik.
def calculate_kodi(data):
    """Kiszámolja a KÖDI értékeket csoportonként.

        Minden csoportban a maximális Ösztöndíjindex 100, a minimális pedig 0.
        A többi hallgató KÖDI értéke ehhez a két értékhez viszonyítva kerül kiszámításra.

        Args:
            data (pd.DataFrame): A hallgatók adatait tartalmazó DataFrame.

        Returns:
            pd.DataFrame: A DataFrame a kiszámolt KÖDI értékekkel.
        """
    grouping_columns = ['KépzésNév', 'Képzési szint', 'Nyelv ID', 'Évfolyam']
    grouped = data.groupby(grouping_columns, observed=True)

    data['MinÖDI'] = grouped['Ösztöndíjindex'].transform('min')
    data['MaxÖDI'] = grouped['Ösztöndíjindex'].transform('max')

    def calculate_group_kodi(row):
        if row['MaxÖDI'] == row['MinÖDI']:
            return 100.0
        else:
            kodi = ((row['Ösztöndíjindex'] - row['MinÖDI']) / (row['MaxÖDI'] - row['MinÖDI'])) * 100
            return round(kodi, 6)

    data['KÖDI'] = data.apply(calculate_group_kodi, axis=1)

    max_odi_mask = data['Ösztöndíjindex'] == data['MaxÖDI']
    data.loc[max_odi_mask, 'KÖDI'] = 100.0

    min_odi_mask = data['Ösztöndíjindex'] == data['MinÖDI']
    data.loc[min_odi_mask, 'KÖDI'] = 0.0

    data.drop(columns=['MinÖDI', 'MaxÖDI'], inplace=True)

    return data




### Streamlit és függvények meghívása

def main():

    st.set_page_config(page_title="Step 1")

    st.title("Student Grouping")
    st.subheader("Upload an input file where 3,8 and 23 are filtered")
    uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx", key="main_file_upload")

    st.subheader("Upload max number of semesters file")
    uploaded_file2 = st.file_uploader("Choose an Excel file", type="xlsx", key="semester_limit_file_upload")


    if uploaded_file is not None and uploaded_file2 is not None:
        data = load_data(uploaded_file)
        semester_limits = load_data(uploaded_file2)
    else:
        st.stop()

    check_duplicate_neptun_codes(data)

    remaining_data, small_groups_data_initial = filter_small_groups(data)
    grouped_data, original_data = group_students(remaining_data)
    updated_data = group_students_by_year(remaining_data)
    updated_data = calculate_scholarship_index(updated_data)
    updated_data = calculate_kodi(updated_data)

    updated_data = remove_lower_kodi_duplicates(updated_data)

    # Re-run the grouping and redistribution after removing duplicates
    remaining_data, small_groups_data_after = filter_small_groups(updated_data)
    grouped_data, original_data = group_students(remaining_data)
    updated_data = group_students_by_year(remaining_data)
    updated_data = calculate_scholarship_index(updated_data)
    updated_data = calculate_kodi(updated_data)

    small_groups_data_combined = (pd.concat([small_groups_data_initial, small_groups_data_after])
                                  .drop_duplicates().reset_index(drop=True))

    small_groups_data_combined = calculate_scholarship_index(small_groups_data_combined)

    updated_data = sort_data(updated_data)
    small_groups_data_combined = sort_data(small_groups_data_combined)

    updated_data = add_group_index(updated_data)
    small_groups_data_combined = add_group_index(small_groups_data_combined)

    updated_data, small_groups_data_combined = highlight_exceeded_semesters(updated_data, small_groups_data_combined,
                                                                            semester_limits)

    main_buffer, separate_buffer = save_to_excel(updated_data, small_groups_data_combined)

    st.subheader("Download Output Files")

    st.download_button(
        label="Download Main Data Excel",
        data=main_buffer.getvalue(),
        file_name="main_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.download_button(
        label="Download Small Groups Data Excel",
        data=separate_buffer.getvalue(),
        file_name="small_groups_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    main()

